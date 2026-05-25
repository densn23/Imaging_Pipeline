from __future__ import annotations

import csv
import math
import pickle
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import DATA_ANALYSIS_ROOT


TABLES_DIR = DATA_ANALYSIS_ROOT / "tables"
STIM_TABLE_CSV = TABLES_DIR / "stim_table_all_jamie.csv"


# -----------------------------------------------------------------------------
# Selection
# -----------------------------------------------------------------------------
MOUSE_NAME = None  # None = all mice, or e.g. "Jamie10" / "Jamie10,Jamie11"


# -----------------------------------------------------------------------------
# Comparison settings
# -----------------------------------------------------------------------------
FREQUENCY_TOL_HZ = 1.0
STIM_DURATION_TOL_S = 0.5
PULSE_WIDTH_TOL_US = 35.0

REFERENCE_PULSES = 10
MAX_PULSE_ANALYSIS_SEC = 0.004
PULSE_PRE_BASELINE_SEC = 0.0005
PULSE_WIDTH_REL_HEIGHT = 0.50


# -----------------------------------------------------------------------------
# Output
# -----------------------------------------------------------------------------
OUTPUT_BASENAME = "stim_label_vs_ephys_audit"
PRINT_ISSUE_ROWS = True
PRINT_STATUSES = ("mismatch", "missing_ephys", "no_pulses", "check", "load_failed")


def parse_name_list(raw: str | None) -> list[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if text == "" or text.lower() in {"none", "all"}:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        out = float(value)
        return out if np.isfinite(out) else None
    text = str(value).strip()
    if not text or text == "?":
        return None
    text = text.replace(",", ".")
    digits = []
    for ch in text:
        if ch.isdigit() or ch in ".-+":
            digits.append(ch)
        elif digits:
            break
    try:
        out = float("".join(digits))
    except Exception:
        return None
    return out if np.isfinite(out) else None


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def trim_float(value: float | None, decimals: int = 3) -> str:
    if value is None or not np.isfinite(value):
        return "?"
    if abs(float(value) - round(float(value))) < 10 ** (-decimals):
        return str(int(round(float(value))))
    return f"{float(value):.{decimals}f}".rstrip("0").rstrip(".")


def mouse_sort_key(mouse: str) -> tuple[int, str]:
    digits = "".join(ch for ch in str(mouse) if ch.isdigit())
    return (int(digits) if digits else 10**9, str(mouse))


def date_sort_key(date_text: str) -> tuple[int, int, int, str]:
    try:
        dt = datetime.strptime(str(date_text), "%d-%m-%y")
        return (dt.year, dt.month, dt.day, date_text)
    except Exception:
        return (9999, 12, 31, str(date_text))


def block_sort_key(block_text: str) -> tuple[int, str]:
    digits = "".join(ch for ch in str(block_text) if ch.isdigit())
    return (int(digits) if digits else 10**9, str(block_text))


def ephys_path_from_row(row: dict[str, str]) -> Path:
    return (
        DATA_ANALYSIS_ROOT
        / row["mouse"]
        / "Open_Ephys"
        / row["date"]
        / row["block"]
        / f"{row['block']}_epoched_ephys.pkl"
    )


def extract_pulse_features(td: dict[str, Any]) -> dict[str, np.ndarray | float] | None:
    stim = np.asarray(td.get("channels", {}).get("stim", []), dtype=float)
    pulse_times = np.asarray(td.get("stim_pulse_times_s", []), dtype=float)
    pulse_idx = np.asarray(td.get("stim_pulse_samples_in_trial", []), dtype=int)
    fs = safe_float(td.get("sample_rate"))

    if stim.size == 0 or pulse_times.size == 0 or pulse_idx.size == 0 or fs is None or fs <= 0:
        return None

    use_n = min(int(REFERENCE_PULSES), int(pulse_idx.size))
    pulse_idx = pulse_idx[:use_n]
    pulse_times = pulse_times[:use_n]

    pre_samp = max(1, int(round(float(PULSE_PRE_BASELINE_SEC) * fs)))
    max_win_samp = max(3, int(round(float(MAX_PULSE_ANALYSIS_SEC) * fs)))

    amps = np.full(use_n, np.nan, dtype=float)
    widths_us = np.full(use_n, np.nan, dtype=float)

    for i, sample_idx in enumerate(pulse_idx):
        j0 = int(sample_idx)
        if j0 < 0 or j0 >= stim.size:
            continue
        next_idx = int(pulse_idx[i + 1]) if i + 1 < use_n else min(stim.size, j0 + max_win_samp)
        j1 = min(stim.size, max(j0 + 1, min(next_idx, j0 + max_win_samp)))

        seg = stim[j0:j1]
        pre = stim[max(0, j0 - pre_samp):j0]
        baseline = float(np.median(pre)) if pre.size else 0.0
        env = np.abs(seg - baseline)
        if env.size == 0:
            continue

        amp = float(np.max(env))
        amps[i] = amp
        if np.isfinite(amp) and amp > 0:
            thr = float(PULSE_WIDTH_REL_HEIGHT) * amp
            active = np.flatnonzero(env >= thr)
            if active.size:
                widths_us[i] = float((active[-1] - active[0] + 1) / fs * 1e6)

    return {
        "amplitude_v": amps,
        "width_us": widths_us,
    }


def summarize_ephys_block(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "status": "missing_ephys",
            "n_trials_total": 0,
            "n_trials_with_pulses": 0,
        }

    try:
        with path.open("rb") as f:
            d = pickle.load(f)
    except Exception as e:
        return {
            "status": "load_failed",
            "load_error": f"{type(e).__name__}: {e}",
            "n_trials_total": 0,
            "n_trials_with_pulses": 0,
        }

    trials = d.get("trials", {})
    freq_vals = []
    dur_vals = []
    count_vals = []
    amp_vals = []
    width_vals = []

    for td in trials.values():
        pulse_times = np.asarray(td.get("stim_pulse_times_s", []), dtype=float)
        pulse_times = pulse_times[np.isfinite(pulse_times)]
        if pulse_times.size == 0:
            continue

        count_vals.append(float(pulse_times.size))
        if pulse_times.size >= 2:
            ipi = np.diff(pulse_times)
            ipi = ipi[np.isfinite(ipi) & (ipi > 0)]
            if ipi.size:
                freq_vals.append(float(1.0 / np.median(ipi)))
                dur_vals.append(float(pulse_times[-1] - pulse_times[0]))

        features = extract_pulse_features(td)
        if features is not None:
            a = np.asarray(features["amplitude_v"], dtype=float)
            a = a[np.isfinite(a)]
            if a.size:
                amp_vals.append(float(np.median(a)))
            w = np.asarray(features["width_us"], dtype=float)
            w = w[np.isfinite(w)]
            if w.size:
                width_vals.append(float(np.median(w)))

    def med(x: list[float]) -> float | None:
        return float(np.nanmedian(x)) if x else None

    def sd(x: list[float]) -> float | None:
        if len(x) < 2:
            return None
        return float(np.nanstd(x, ddof=1))

    n_with = len(count_vals)
    if n_with == 0:
        return {
            "status": "no_pulses",
            "n_trials_total": len(trials),
            "n_trials_with_pulses": 0,
        }

    return {
        "status": "ok",
        "n_trials_total": len(trials),
        "n_trials_with_pulses": n_with,
        "freq_hz_median": med(freq_vals),
        "freq_hz_sd": sd(freq_vals),
        "stim_duration_s_median": med(dur_vals),
        "stim_duration_s_sd": sd(dur_vals),
        "pulse_count_median": med(count_vals),
        "pulse_count_sd": sd(count_vals),
        "amplitude_v_median": med(amp_vals),
        "amplitude_v_sd": sd(amp_vals),
        "pulse_width_us_median": med(width_vals),
        "pulse_width_us_sd": sd(width_vals),
    }


def compare_numeric(label_value: float | None, actual_value: float | None, tol: float) -> tuple[str, float | None]:
    if label_value is None:
        return ("unchecked", None)
    if actual_value is None or not np.isfinite(actual_value):
        return ("missing_actual", None)
    diff = float(actual_value - label_value)
    if abs(diff) <= float(tol):
        return ("ok", diff)
    return ("mismatch", diff)


def compare_row(row: dict[str, str], metrics: dict[str, Any], ephys_path: Path) -> dict[str, Any]:
    label_freq = safe_float(row.get("frequency_hz"))
    label_pw = safe_float(row.get("pulse_width_us"))
    label_pw_trace = (2.0 * label_pw) if label_pw is not None else None
    label_stim = safe_float(row.get("stimulation_time_s"))

    freq_status, freq_diff = compare_numeric(label_freq, metrics.get("freq_hz_median"), FREQUENCY_TOL_HZ)
    pw_status, pw_diff = compare_numeric(label_pw_trace, metrics.get("pulse_width_us_median"), PULSE_WIDTH_TOL_US)
    stim_status, stim_diff = compare_numeric(label_stim, metrics.get("stim_duration_s_median"), STIM_DURATION_TOL_S)

    reasons: list[str] = []
    if metrics["status"] == "missing_ephys":
        overall = "missing_ephys"
        reasons.append("missing ephys file")
    elif metrics["status"] == "load_failed":
        overall = "load_failed"
        reasons.append(metrics.get("load_error", "load failed"))
    elif metrics["status"] == "no_pulses":
        overall = "no_pulses"
        reasons.append("no stim pulses found")
    else:
        for name, status in [
            ("frequency", freq_status),
            ("pulse_width", pw_status),
            ("stim_duration", stim_status),
        ]:
            if status == "mismatch":
                reasons.append(name)
            elif status == "missing_actual":
                reasons.append(f"{name}_missing_actual")

        if any(status == "mismatch" for status in [freq_status, pw_status, stim_status]):
            overall = "mismatch"
        elif any(status == "missing_actual" for status in [freq_status, pw_status, stim_status]):
            overall = "check"
        else:
            overall = "ok"

    return {
        "mouse": row.get("mouse", "?"),
        "phase": row.get("phase", "?"),
        "date": row.get("date", "?"),
        "block": row.get("block", "?"),
        "protocol_raw": row.get("protocol_raw", "?"),
        "imaging_side": row.get("imaging_side", "?"),
        "label_frequency_hz": label_freq,
        "label_amplitude_uA": row.get("amplitude_uA", "?"),
        "label_pulse_width_us": label_pw,
        "label_pulse_width_us_trace_expected": label_pw_trace,
        "label_stim_duration_s": label_stim,
        "ephys_n_trials_total": metrics.get("n_trials_total"),
        "ephys_n_trials_with_pulses": metrics.get("n_trials_with_pulses"),
        "ephys_frequency_hz": metrics.get("freq_hz_median"),
        "ephys_frequency_hz_sd": metrics.get("freq_hz_sd"),
        "frequency_status": freq_status,
        "frequency_diff_hz": freq_diff,
        "ephys_pulse_width_us": metrics.get("pulse_width_us_median"),
        "ephys_pulse_width_us_sd": metrics.get("pulse_width_us_sd"),
        "pulse_width_status": pw_status,
        "pulse_width_diff_us": pw_diff,
        "ephys_stim_duration_s": metrics.get("stim_duration_s_median"),
        "ephys_stim_duration_s_sd": metrics.get("stim_duration_s_sd"),
        "stim_duration_status": stim_status,
        "stim_duration_diff_s": stim_diff,
        "ephys_pulse_count": metrics.get("pulse_count_median"),
        "ephys_pulse_count_sd": metrics.get("pulse_count_sd"),
        "ephys_amplitude_v": metrics.get("amplitude_v_median"),
        "ephys_amplitude_v_sd": metrics.get("amplitude_v_sd"),
        "overall_status": overall,
        "mismatch_reasons": ", ".join(reasons) if reasons else "",
        "ephys_path": str(ephys_path),
    }


def load_rows() -> list[dict[str, str]]:
    mice = set(parse_name_list(MOUSE_NAME))
    rows = []
    with STIM_TABLE_CSV.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if mice and row.get("mouse", "") not in mice:
                continue
            rows.append(row)
    rows.sort(key=lambda r: (mouse_sort_key(r.get("mouse", "")), date_sort_key(r.get("date", "")), block_sort_key(r.get("block", ""))))
    return rows


def build_audit_rows(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    cache: dict[Path, dict[str, Any]] = {}
    out = []
    for row in rows:
        path = ephys_path_from_row(row)
        if path not in cache:
            cache[path] = summarize_ephys_block(path)
        out.append(compare_row(row, cache[path], path))
    return out


def slugify(text: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in str(text))
    return cleaned.strip("_") or "audit"


def output_path() -> Path:
    mice = parse_name_list(MOUSE_NAME)
    if mice:
        suffix = "_" + "_".join(slugify(m) for m in mice)
    else:
        suffix = "_all_mice"
    return TABLES_DIR / f"{OUTPUT_BASENAME}{suffix}.xlsx"


def autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        lengths = []
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = cell.value
            if val is None:
                continue
            lengths.append(len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(lengths, default=8) + 2, 45)


def write_sheet(ws, rows: list[dict[str, Any]], headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, header in enumerate(headers, start=1):
        cell = ws.cell(1, c, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_fills = {
        "ok": PatternFill("solid", fgColor="E2F0D9"),
        "mismatch": PatternFill("solid", fgColor="FCE4D6"),
        "check": PatternFill("solid", fgColor="FFF2CC"),
        "missing_ephys": PatternFill("solid", fgColor="F4CCCC"),
        "load_failed": PatternFill("solid", fgColor="F4CCCC"),
        "no_pulses": PatternFill("solid", fgColor="F4CCCC"),
    }

    for r_idx, row in enumerate(rows, start=2):
        status_fill = status_fills.get(str(row.get("overall_status", "")).strip(), PatternFill(fill_type=None))
        for c_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", horizontal="center" if isinstance(value, (int, float)) else "left")
            if header == "overall_status":
                cell.fill = status_fill
                cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).column_letter}{max(2, len(rows) + 1)}"
    autosize_columns(ws)


def save_workbook(audit_rows: list[dict[str, Any]], out_xlsx: Path) -> None:
    wb = Workbook()

    all_headers = [
        "mouse",
        "phase",
        "date",
        "block",
        "protocol_raw",
        "imaging_side",
        "label_frequency_hz",
        "label_amplitude_uA",
        "label_pulse_width_us",
        "label_pulse_width_us_trace_expected",
        "label_stim_duration_s",
        "ephys_n_trials_total",
        "ephys_n_trials_with_pulses",
        "ephys_frequency_hz",
        "ephys_frequency_hz_sd",
        "frequency_status",
        "frequency_diff_hz",
        "ephys_pulse_width_us",
        "ephys_pulse_width_us_sd",
        "pulse_width_status",
        "pulse_width_diff_us",
        "ephys_stim_duration_s",
        "ephys_stim_duration_s_sd",
        "stim_duration_status",
        "stim_duration_diff_s",
        "ephys_pulse_count",
        "ephys_pulse_count_sd",
        "ephys_amplitude_v",
        "ephys_amplitude_v_sd",
        "overall_status",
        "mismatch_reasons",
        "ephys_path",
    ]

    mismatches = [r for r in audit_rows if str(r.get("overall_status")) != "ok"]
    summary_counts = Counter(str(r.get("overall_status", "?")) for r in audit_rows)

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Stim label vs ephys audit"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Mouse filter"
    ws_summary["B3"] = ", ".join(parse_name_list(MOUSE_NAME)) if parse_name_list(MOUSE_NAME) else "All mice"
    ws_summary["A4"] = "Rows audited"
    ws_summary["B4"] = len(audit_rows)
    ws_summary["A6"] = "Overall status"
    ws_summary["B6"] = "Count"

    r = 7
    for status in ["ok", "mismatch", "check", "missing_ephys", "load_failed", "no_pulses"]:
        ws_summary.cell(r, 1, status)
        ws_summary.cell(r, 2, int(summary_counts.get(status, 0)))
        r += 1

    ws_summary["D3"] = "Frequency tol (Hz)"
    ws_summary["E3"] = FREQUENCY_TOL_HZ
    ws_summary["D4"] = "Stim duration tol (s)"
    ws_summary["E4"] = STIM_DURATION_TOL_S
    ws_summary["D5"] = "Pulse width tol (us)"
    ws_summary["E5"] = PULSE_WIDTH_TOL_US
    autosize_columns(ws_summary)

    ws_mismatch = wb.create_sheet("Mismatches")
    write_sheet(ws_mismatch, mismatches, all_headers)

    ws_all = wb.create_sheet("All Rows")
    write_sheet(ws_all, audit_rows, all_headers)

    wb.save(out_xlsx)


def print_summary(audit_rows: list[dict[str, Any]], out_xlsx: Path) -> None:
    counts = Counter(str(r.get("overall_status", "?")) for r in audit_rows)
    print(f"[SAVED] {out_xlsx}")
    print(f"[ROWS] {len(audit_rows)}")
    for status in ["ok", "mismatch", "check", "missing_ephys", "load_failed", "no_pulses"]:
        print(f"[{status.upper()}] {int(counts.get(status, 0))}")


def print_issue_rows(audit_rows: list[dict[str, Any]]) -> None:
    def sort_key(row: dict[str, Any]) -> tuple[tuple[int, str], tuple[int, int, int, str], tuple[int, str]]:
        return (
            mouse_sort_key(str(row.get("mouse", ""))),
            date_sort_key(str(row.get("date", ""))),
            block_sort_key(str(row.get("block", ""))),
        )

    by_status: dict[str, list[dict[str, Any]]] = {}
    for status in PRINT_STATUSES:
        rows = [r for r in audit_rows if str(r.get("overall_status", "")) == status]
        if rows:
            by_status[status] = sorted(rows, key=sort_key)

    for status in PRINT_STATUSES:
        rows = by_status.get(status, [])
        if not rows:
            continue
        print(f"\n[{status.upper()} ROWS]")
        for row in rows:
            label_f = trim_float(safe_float(row.get("label_frequency_hz")))
            label_pw = trim_float(safe_float(row.get("label_pulse_width_us")))
            label_stim = trim_float(safe_float(row.get("label_stim_duration_s")))
            ephys_f = trim_float(safe_float(row.get("ephys_frequency_hz")))
            ephys_pw = trim_float(safe_float(row.get("ephys_pulse_width_us")))
            ephys_stim = trim_float(safe_float(row.get("ephys_stim_duration_s")))
            reasons = normalize_text(row.get("mismatch_reasons", "")) or "?"
            print(
                f"  {row.get('mouse','?')} | {row.get('date','?')} | {row.get('block','?')} | "
                f"{reasons} | label f/pw/stim={label_f}/{label_pw}/{label_stim} | "
                f"ephys f/pw/stim={ephys_f}/{ephys_pw}/{ephys_stim}"
            )


def main() -> None:
    rows = load_rows()
    audit_rows = build_audit_rows(rows)
    out_xlsx = output_path()
    save_workbook(audit_rows, out_xlsx)
    print_summary(audit_rows, out_xlsx)
    if PRINT_ISSUE_ROWS:
        print_issue_rows(audit_rows)


if __name__ == "__main__":
    main()

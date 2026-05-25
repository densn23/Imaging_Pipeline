from __future__ import annotations

import csv
import math
import re
import zipfile
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
import xml.etree.ElementTree as ET

import pickle
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import DATA_ANALYSIS_ROOT


TABLES_DIR = DATA_ANALYSIS_ROOT / "tables"
WORKBOOK_PATH = TABLES_DIR / "Experimental logbook - Maxime.xlsx"
OUTPUT_CSV = TABLES_DIR / "stim_table_all_jamie.csv"
SUMMARY_XLSX = TABLES_DIR / "stim_table_summary_grouped_textdates.xlsx"

MICE = ["Jamie6", "Jamie8", "Jamie10", "Jamie11", "Jamie12", "Vinnie1", "Vinnie2"]
FIBER_SITE = {
    "Jamie6": "STN",
    "Jamie8": "ZI",
    "Jamie10": "STN",
    "Jamie11": "STN",
    "Jamie12": "STN",
    "Vinnie1": "STN",
    "Vinnie2": "STN",
}
DEFAULT_THR_UA = {
    "Jamie8": 25.0,
    "Jamie10": 25.0,
    "Jamie11": 25.0,
    "Jamie12": 30.0,
    "Vinnie1": 25.0,
    "Vinnie2": 25.0,
}
IMAGING_SIDE_OVERRIDES = {
}
VINNIE1_LEFT_IMAGED_RIGHT_DBS_BLOCKS = {
    ("Vinnie1", "11-05-26", "R18"),
    ("Vinnie1", "12-05-26", "R5"),
    ("Vinnie1", "12-05-26", "R6"),
    ("Vinnie1", "12-05-26", "R7"),
    ("Vinnie1", "12-05-26", "R14"),
    ("Vinnie1", "13-05-26", "R20"),
    ("Vinnie1", "20-05-26", "R6"),
}
SUMMARY_MOUSE_ORDER = ["Jamie6", "Jamie8", "Jamie10", "Jamie11", "Jamie12", "Vinnie1", "Vinnie2"]

NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

BASE_TEED_AMP_UA = 25.0
BASE_TEED_FREQ_HZ = 135.0
BASE_TEED_PW_US = 100.0
TEED_CONSTANT = BASE_TEED_AMP_UA ** 2 * BASE_TEED_FREQ_HZ * BASE_TEED_PW_US


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip())


def col_to_num(col: str) -> int:
    n = 0
    for c in col:
        if c.isalpha():
            n = n * 26 + (ord(c.upper()) - 64)
    return n


def excel_serial_to_date_text(value: str) -> str:
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return "?"
    dt = datetime(1899, 12, 30) + timedelta(days=serial)
    return dt.strftime("%d-%m-%y")


def format_number(value: float, ndigits: int = 3) -> str:
    if not np.isfinite(value):
        return "?"
    rounded = round(float(value), ndigits)
    if abs(rounded - round(rounded)) < 10 ** (-ndigits):
        return str(int(round(rounded)))
    return f"{rounded:.{ndigits}f}".rstrip("0").rstrip(".")


def normalize_numeric_text(value: str) -> str:
    text = normalize_text(value)
    if text in {"", "?"}:
        return "?" if text == "?" else text
    try:
        return format_number(float(text), ndigits=3)
    except (TypeError, ValueError):
        return text


def block_sort_value(block: str) -> tuple[int, str]:
    m = re.search(r"R(\d+)", str(block or ""))
    return (int(m.group(1)), str(block)) if m else (10**9, str(block))


def parse_date_for_sort(date_text: str) -> tuple[int, int, int]:
    try:
        dt = datetime.strptime(str(date_text), "%d-%m-%y")
        return (dt.year, dt.month, dt.day)
    except Exception:
        return (9999, 12, 31)


def extract_blocks(block_text: str) -> list[str]:
    return [f"R{m}" for m in re.findall(r"[Rr]\s*(\d+)", block_text or "")]


def parse_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    except KeyError:
        return []

    values = []
    for si in root.findall("a:si", NS):
        parts = []
        for t in si.iterfind(".//a:t", NS):
            parts.append(t.text or "")
        values.append("".join(parts))
    return values


def workbook_sheet_targets(zf: zipfile.ZipFile) -> dict[str, str]:
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    out = {}
    for sheet in wb.find("a:sheets", NS):
        name = sheet.attrib.get("name", "")
        rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = relmap.get(rid)
        if name and target:
            out[name] = target
    return out


def read_sheet_rows(zf: zipfile.ZipFile, target: str, shared: list[str]) -> list[dict[int, str]]:
    root = ET.fromstring(zf.read("xl/" + target))
    rows = []
    for row in root.findall(".//a:sheetData/a:row", NS):
        vals: dict[int, str] = {-1: row.attrib.get("r", "?")}
        for cell in row.findall("a:c", NS):
            ref = cell.attrib.get("r", "")
            col = "".join(ch for ch in ref if ch.isalpha())
            cnum = col_to_num(col)
            typ = cell.attrib.get("t")
            val = cell.find("a:v", NS)

            if typ == "s" and val is not None and val.text is not None:
                text = shared[int(val.text)]
            elif typ == "inlineStr":
                text = "".join(t.text or "" for t in cell.findall(".//a:t", NS))
            elif val is not None:
                text = val.text or ""
            else:
                text = "".join(t.text or "" for t in cell.findall(".//a:t", NS))

            text = normalize_text(text)
            if text:
                vals[cnum] = text
        if vals:
            rows.append(vals)
    return rows


def infer_exposure_ms(mouse: str, date: str, block: str) -> str:
    summary_path = DATA_ANALYSIS_ROOT / mouse / "Imaging_Data" / date / block / f"{block}_summary.pkl"
    if summary_path.exists():
        try:
            with open(summary_path, "rb") as f:
                summary = pickle.load(f)
            fps = summary.get("summary", {}).get("ephys", {}).get("cam_fps_hz_mean")
            if fps is None:
                fps = summary.get("summary", {}).get("processed_notched", {}).get("fps_hz_mean")
            fps = float(fps)
            if np.isfinite(fps):
                if 800 <= fps <= 1100:
                    return "1"
                if 400 <= fps <= 600:
                    return "2"
                return format_number(1000.0 / fps, ndigits=2)
        except Exception:
            pass

    ephys_path = DATA_ANALYSIS_ROOT / mouse / "Open_Ephys" / date / block / f"{block}_epoched_ephys.pkl"
    if ephys_path.exists():
        try:
            with open(ephys_path, "rb") as f:
                eph = pickle.load(f)
            fps_vals = []
            for td in eph.get("trials", {}).values():
                fps = td.get("cam_fps_hz")
                try:
                    fps = float(fps)
                except (TypeError, ValueError):
                    continue
                if np.isfinite(fps) and fps > 0:
                    fps_vals.append(fps)
            if fps_vals:
                fps = float(np.nanmedian(fps_vals))
                if 800 <= fps <= 1100:
                    return "1"
                if 400 <= fps <= 600:
                    return "2"
                return format_number(1000.0 / fps, ndigits=2)
        except Exception:
            pass

    return "?"


def infer_exposure_ms_from_trace_count(mouse: str, date: str, block: str, stim_s: str) -> str:
    try:
        duration_s = float(stim_s)
    except (TypeError, ValueError):
        return "?"
    if not np.isfinite(duration_s) or duration_s <= 0:
        return "?"

    traces_path = DATA_ANALYSIS_ROOT / mouse / "Imaging_Data" / date / block / f"{block}_traces.pkl"
    if not traces_path.exists():
        return "?"

    try:
        with open(traces_path, "rb") as f:
            traces = pickle.load(f)
    except Exception:
        return "?"

    n_frames = []
    for td in traces.get("trials", {}).values():
        n = td.get("n_frames")
        if n is None and hasattr(td.get("trace_raw"), "__len__"):
            n = len(td["trace_raw"])
        try:
            n = float(n)
        except (TypeError, ValueError):
            continue
        if np.isfinite(n) and n > 0:
            n_frames.append(n)

    if not n_frames:
        return "?"

    fps = float(np.nanmedian(n_frames)) / duration_s
    if 800 <= fps <= 1100:
        return "1"
    if 400 <= fps <= 600:
        return "2"
    if fps > 0:
        return format_number(1000.0 / fps, ndigits=2)
    return "?"


def infer_stim_duration_s(mouse: str, date: str, block: str) -> str:
    ephys_path = DATA_ANALYSIS_ROOT / mouse / "Open_Ephys" / date / block / f"{block}_epoched_ephys.pkl"
    if not ephys_path.exists():
        return "?"

    try:
        with open(ephys_path, "rb") as f:
            eph = pickle.load(f)
    except Exception:
        return "?"

    durations = []
    for td in eph.get("trials", {}).values():
        pulses = np.asarray(td.get("stim_pulse_times_s", []), dtype=float)
        pulses = pulses[np.isfinite(pulses)]
        if len(pulses) >= 2:
            durations.append(float(pulses[-1] - pulses[0]))

    if not durations:
        return "?"

    dur = float(np.nanmedian(durations))
    if not np.isfinite(dur):
        return "?"
    return format_number(dur, ndigits=2)


def infer_teed_amplitude(freq_hz: float, pw_us: float) -> str:
    if not (np.isfinite(freq_hz) and freq_hz > 0 and np.isfinite(pw_us) and pw_us > 0):
        return "?"
    amp = math.sqrt(TEED_CONSTANT / (float(freq_hz) * float(pw_us)))
    return format_number(amp, ndigits=2)


def infer_default_exposure_from_text(entry: dict[str, str]) -> str:
    text = normalize_text(entry.get('protocol_raw', '')).lower()
    if any(token in text for token in ["1khz", "1 khz", "1000 fps", "fps 1000", "exp=1", "exp = 1", "spike triggered"]):
        return "1"
    if entry.get("mouse") in {"Jamie11", "Jamie12"}:
        return "2"
    return "?"


def infer_imaging_side_from_text(text: str) -> str:
    low = normalize_text(text).lower()
    matches = re.findall(r"\b(left|right)\b", low)
    if matches:
        side = matches[-1]
        return "Left" if side == "left" else "Right"
    return "Left"


def apply_vinnie1_imaging_side(entry: dict[str, str]) -> None:
    if entry.get("mouse") != "Vinnie1":
        return
    key = (entry.get("mouse", ""), entry.get("date", ""), entry.get("block", ""))
    entry["imaging_side"] = "Left" if key in VINNIE1_LEFT_IMAGED_RIGHT_DBS_BLOCKS else "Right"


def infer_special_amplitude_category(entry: dict[str, str]) -> str | None:
    text = normalize_text(
        f"{entry.get('protocol', '')} {entry.get('protocol_raw', '')}"
    ).lower()
    if "thetaburst" in text:
        return "theta"
    if "ramp" in text or "ramping" in text:
        return "ramp"
    return None


def normalize_phase_label(value: str) -> str:
    low = normalize_text(value).lower()
    if not low or low == "?":
        return "?"
    if "pre" in low and "6ohda" in low:
        return "Pre"
    if any(token in low for token in ["post", "acute", "chronic"]) and "6ohda" in low:
        return "Post"
    return "?"


def is_mixed_cortex_entry(entry: dict[str, str]) -> bool:
    text = normalize_text(entry.get("protocol_raw", "")).lower()
    return ("m1-dbs" in text) or ("stn fiber" in text)


def is_jamie8_tail_baseline(entry: dict[str, str]) -> bool:
    return (
        entry.get("mouse") == "Jamie8"
        and entry.get("date") == "12-01-25"
        and entry.get("block") in {"R2", "R3"}
    )


def parse_explicit_numeric_triplet(desc: str) -> tuple[str, str, str, str]:
    m = re.search(
        r"\(\s*(\d+(?:\.\d+)?)\s*,\s*(\d+(?:\.\d+)?)\s*,\s*(\d+(?:\.\d+)?)\s*,\s*(\d+(?:\.\d+)?)\s*sec",
        desc,
        flags=re.IGNORECASE,
    )
    if not m:
        return "?", "?", "?", "?"
    freq = format_number(float(m.group(1)), ndigits=1)
    amp = format_number(float(m.group(2)), ndigits=2)
    pw = format_number(float(m.group(3)), ndigits=1)
    stim_s = format_number(float(m.group(4)), ndigits=2)
    return freq, amp, pw, stim_s


def parse_protocol(desc: str) -> str:
    low = desc.lower()
    if "pulsewidth analysis" in low:
        direction = "ramp up" if "ramp up" in low else "ramp down" if "ramp down" in low else ""
        return normalize_text(f"pulsewidth analysis {direction}".strip())
    if "amplitude ramping" in low:
        return "amplitude ramping"
    if "frequency ramping" in low:
        return "frequency ramping"
    if "spike triggered" in low:
        return "spike triggered"
    if "thetaburst" in low:
        return "thetaburst"
    if "low pulse width" in low:
        return "low pulse width"
    if "high pulse width" in low:
        return "high pulse width"
    if "low frequency" in low:
        return "low frequency"
    if "high frequency" in low:
        return "high frequency"
    if "baseline below thr" in low:
        return "baseline below thr"
    if "baseline above thr" in low:
        return "baseline above thr"
    if "baseline at thr" in low:
        return "baseline at thr"
    if low.startswith("baseline"):
        return "baseline"
    if "fluorescence test" in low:
        return "fluorescence test"
    return normalize_text(desc.split("–")[0].split("-")[0])


def parse_frequency(desc: str, protocol: str) -> tuple[str, str]:
    low = desc.lower()
    explicit_freq, _, _, _ = parse_explicit_numeric_triplet(desc)
    if explicit_freq != "?":
        return explicit_freq, explicit_freq
    if "frequency ramping" in low:
        before_hz = re.split(r"hz", desc, maxsplit=1, flags=re.IGNORECASE)[0]
        nums = re.findall(r"\d+(?:\.\d+)?", before_hz)
        if len(nums) >= 2:
            values = [format_number(float(n), ndigits=1) for n in nums]
            return "-".join(values), "-".join(values)
        m = re.search(r"(\d+(?:-\d+)+)\s*hz", desc, flags=re.IGNORECASE)
        if m:
            return m.group(1), m.group(1)
        return "?", "frequency ramping"

    m = re.search(r"(\d+(?:\.\d+)?)\s*hz", desc, flags=re.IGNORECASE)
    if m:
        val = format_number(float(m.group(1)), ndigits=1)
        return val, val

    return "?", "?"


def parse_pulse_width(desc: str, protocol: str) -> tuple[str, str]:
    _, _, explicit_pw, _ = parse_explicit_numeric_triplet(desc)
    if explicit_pw != "?":
        return explicit_pw, explicit_pw
    if "pulsewidth analysis" in desc.lower():
        return "?", "pulse width ramp"

    m = re.search(r"(\d+(?:\.\d+)?)\s*[µμu]\s*s", desc, flags=re.IGNORECASE)
    if m:
        val = format_number(float(m.group(1)), ndigits=1)
        return val, val

    m = re.search(r"(\d+(?:\.\d+)?)\s*p\s*w\b", desc, flags=re.IGNORECASE)
    if m:
        val = format_number(float(m.group(1)), ndigits=1)
        return val, val

    return "?", "?"


def parse_stim_duration(desc: str, protocol: str) -> tuple[str, str]:
    low = desc.lower()
    if protocol == "baseline":
        return "?", "baseline"
    if "pulsewidth analysis" in low:
        if "1s per parameter" in low:
            return "?", "1s per parameter"
        return "?", "?"
    if "amplitude ramping" in low or "frequency ramping" in low:
        return "?", "ramping"

    freq, amp, pw, explicit = parse_explicit_numeric_triplet(desc)
    if explicit != "?":
        return explicit, explicit

    m = re.search(r",\s*(\d+(?:\.\d+)?)\s*(sec|s|min)\b", desc, flags=re.IGNORECASE)
    if m:
        val = float(m.group(1))
        unit = m.group(2).lower()
        seconds = val * 60.0 if unit == "min" else val
        text = f"{val:g}{unit}"
        return format_number(seconds, ndigits=2), text

    m = re.search(r"\(\s*\d+\s*x\s*(\d+(?:\.\d+)?)\s*(s|min)\s*\)", desc, flags=re.IGNORECASE)
    if m:
        val = float(m.group(1))
        unit = m.group(2).lower()
        seconds = val * 60.0 if unit == "min" else val
        text = f"{val:g}{unit}"
        return format_number(seconds, ndigits=2), text

    all_durations = re.findall(r"(\d+(?:\.\d+)?)\s*(sec|s|min)\b", desc, flags=re.IGNORECASE)
    if all_durations:
        val = float(all_durations[-1][0])
        unit = all_durations[-1][1].lower()
        seconds = val * 60.0 if unit == "min" else val
        text = f"{val:g}{unit}"
        return format_number(seconds, ndigits=2), text

    return "?", "?"


def parse_exposure(text: str) -> str:
    m = re.search(r"exp\s*=\s*(\d+(?:\.\d+)?)", text, flags=re.IGNORECASE)
    if m:
        return format_number(float(m.group(1)), ndigits=2)

    m = re.search(r"exposure\s*(?:=|at)?\s*(\d+(?:\.\d+)?)", text, flags=re.IGNORECASE)
    if m:
        return format_number(float(m.group(1)), ndigits=2)

    if re.search(r"\b4\s*k\s*hz\b|\b4khz\b", text, flags=re.IGNORECASE):
        return "0.25"
    if re.search(r"\b2\s*k\s*hz\b|\b2khz\b", text, flags=re.IGNORECASE):
        return "0.5"
    if re.search(r"\b1000\s*fps\b|\b1\s*k\s*hz\b|\b1khz\b", text, flags=re.IGNORECASE):
        return "1"
    if re.search(r"\b500\s*fps\b|\b500\s*hz\b", text, flags=re.IGNORECASE):
        return "2"

    return "?"


def parse_amplitude(text: str, protocol: str, mouse: str, freq_hz: str, pw_us: str) -> tuple[str, str]:
    low = text.lower()
    thr = DEFAULT_THR_UA.get(mouse, np.nan)

    if "amplitude ramping" in low:
        return "?", "amplitude ramping"
    if "frequency ramping" in low:
        return "?", "TEED-balanced across frequency"
    if "pulsewidth analysis" in low:
        return "?", "TEED-balanced across pulse widths"

    _, explicit_amp, _, _ = parse_explicit_numeric_triplet(text)
    if explicit_amp != "?":
        return explicit_amp, explicit_amp

    m = re.search(r"thr\s*\*\s*\(?\s*(\d+(?:\.\d+)?)\s*\)?", low)
    if m and np.isfinite(thr):
        factor = float(m.group(1))
        val = format_number(thr * factor, ndigits=2)
        return val, f"thr*{factor:g}"

    m = re.search(r"\((\d+(?:\.\d+)?)\s*u?a\)", low)
    if m:
        val = format_number(float(m.group(1)), ndigits=2)
        return val, val

    m = re.search(r"amp(?:litude)?\s*(?:=|at)?\s*(\d+(?:\.\d+)?)", low, flags=re.IGNORECASE)
    if m and "thr" not in low[m.start() : m.end() + 8]:
        val = format_number(float(m.group(1)), ndigits=2)
        return val, val

    m = re.search(r"(\d+(?:\.\d+)?)\s*u\s*a\b", low, flags=re.IGNORECASE)
    if m:
        val = format_number(float(m.group(1)), ndigits=2)
        return val, val

    m = re.search(r"(\d+(?:\.\d+)?)\s*amp\b", low, flags=re.IGNORECASE)
    if m:
        val = format_number(float(m.group(1)), ndigits=2)
        return val, val

    m = re.search(r"thr\s*\*\s*\(?\s*(\d+(?:\.\d+)?)\s*\)?", low)
    if m and np.isfinite(thr):
        factor = float(m.group(1))
        val = format_number(thr * factor, ndigits=2)
        return val, f"thr*{factor:g}"

    if "above thr" in low:
        m = re.search(r"\((\d+(?:\.\d+)?)\s*u?a\)", low)
        if m:
            val = format_number(float(m.group(1)), ndigits=2)
            return val, f"above thr ({val})"

    if "amp at thr" in low or "amp thr" in low or "at thr" in low:
        if np.isfinite(thr):
            val = format_number(thr, ndigits=2)
            return val, "thr"
        return "?", "thr"

    if "teed" in low:
        try:
            freq_val = float(freq_hz)
            pw_val = float(pw_us)
        except (TypeError, ValueError):
            return "?", "TEED"
        return infer_teed_amplitude(freq_val, pw_val), "TEED"

    if re.fullmatch(r"\d+(?:\.\d+)?", low.strip()):
        val = format_number(float(low.strip()), ndigits=2)
        return val, val

    return "?", "?"


def parse_jamie_wide_description(desc: str, notes: str, mouse: str) -> dict[str, str]:
    desc = normalize_text(desc)
    notes = normalize_text(notes)
    combined = normalize_text(f"{desc} {notes}")
    protocol = parse_protocol(desc)
    freq_hz, freq_raw = parse_frequency(desc, protocol)
    pw_us, pw_raw = parse_pulse_width(desc, protocol)
    stim_s, stim_raw = parse_stim_duration(desc, protocol)
    exposure_ms = parse_exposure(combined)
    amp_uA, amp_raw = parse_amplitude(combined, protocol, mouse, freq_hz, pw_us)

    dbs_side = "right" if "right-dbs" in desc.lower() else "left" if "left-dbs" in desc.lower() else "left"
    fiber_placement = f"{FIBER_SITE[mouse]} {('left' if mouse in ('Jamie8', 'Jamie10') else '?')}"
    dbs_placement = f"DBS {dbs_side}"
    placement_summary = f"Fiber: {fiber_placement} | {dbs_placement}"
    imaging_side = infer_imaging_side_from_text(desc)

    return {
        "protocol": protocol,
        "protocol_raw": desc,
        "frequency_hz": freq_hz,
        "frequency_raw": freq_raw,
        "amplitude_uA": amp_uA,
        "amplitude_raw": amp_raw,
        "pulse_width_us": pw_us,
        "pulse_width_raw": pw_raw,
        "exposure_ms": exposure_ms,
        "stimulation_time_s": stim_s,
        "stimulation_time_raw": stim_raw,
        "fiber_placement": fiber_placement,
        "dbs_placement": dbs_placement,
        "placement_summary": placement_summary,
        "imaging_side": imaging_side,
        "notes": normalize_text(notes),
    }


def parse_jamie6_rows(rows: list[dict[int, str]]) -> list[dict[str, str]]:
    out = []
    current_phase = "?"

    for row in rows:
        c1 = normalize_text(row.get(1, ""))
        c2 = normalize_text(row.get(2, ""))
        c3 = normalize_text(row.get(3, ""))
        c4 = normalize_text(row.get(4, ""))

        low1 = c1.lower()
        if "pre-6ohda" in low1:
            current_phase = "Pre-6OHDA"
            continue
        if "post-6ohda" in low1:
            current_phase = "Post-6OHDA"
            continue

        blocks = extract_blocks(c2)
        if not blocks:
            continue

        if normalize_text(c3).lower() != "stn ipsi" or normalize_text(c4).lower() != "stn ipsi":
            continue

        date_text = excel_serial_to_date_text(c1)
        freq = normalize_text(row.get(6, "")) or "?"
        exp = normalize_text(row.get(7, "")) or "?"
        stim = normalize_text(row.get(8, "")) or "?"
        pw = normalize_text(row.get(9, "")) or "?"
        amp = normalize_text(row.get(10, "")) or "?"
        extra = " | ".join(
            normalize_text(v)
            for k, v in sorted(row.items())
            if k >= 11 and normalize_text(v)
        )

        protocol = "standard"
        low_extra = extra.lower()
        if "-" in freq and any(ch.isdigit() for ch in freq):
            protocol = "frequency ramping"
        elif "thetaburst" in low_extra:
            protocol = "thetaburst"
        elif "baseline" in low_extra:
            protocol = "baseline"

        for block in blocks:
            out.append(
                {
                    "mouse": "Jamie6",
                    "phase": current_phase,
                    "date": date_text,
                    "session": "?",
                    "block": block,
                    "protocol": protocol,
                    "protocol_raw": extra if extra else protocol,
                    "fiber_placement": "STN ipsi",
                    "dbs_placement": "STN ipsi",
                    "placement_summary": "Fiber: STN ipsi | DBS: STN ipsi",
                    "imaging_side": infer_imaging_side_from_text(extra),
                    "frequency_hz": freq,
                    "frequency_raw": freq,
                    "amplitude_uA": amp,
                    "amplitude_raw": amp,
                    "pulse_width_us": pw,
                    "pulse_width_raw": pw,
                    "exposure_ms": exp,
                    "stimulation_time_s": stim,
                    "stimulation_time_raw": stim,
                    "notes": extra if extra else "?",
                    "source_sheet": "Jamie6",
                    "source_row": str(row.get(-1, "?")),
                }
            )

    return out


def parse_jamie_wide_rows(rows: list[dict[int, str]], mouse: str) -> list[dict[str, str]]:
    out = []
    current_phase = "?"
    current_date = "?"
    current_session = "?"

    for row in rows:
        c1 = normalize_text(row.get(1, ""))
        c2 = normalize_text(row.get(2, ""))
        c3 = normalize_text(row.get(3, ""))

        low1 = c1.lower()
        if "pre-6ohda" in low1:
            current_phase = "Pre-6OHDA"
        elif "acute-post-6ohda" in low1:
            current_phase = "Acute-Post-6OHDA"
        elif "post-6ohda" in low1:
            current_phase = "Post-6OHDA"

        try:
            if c2:
                current_date = excel_serial_to_date_text(c2)
        except Exception:
            pass
        if mouse == "Vinnie1" and str(row.get(-1, "")) == "12":
            current_date = "15-05-26"
        if c3:
            current_session = c3

        max_col = max((k for k in row.keys() if k > 0), default=0)
        for desc_col in range(4, max_col + 1, 3):
            block_col, note_col = desc_col + 1, desc_col + 2
            desc = normalize_text(row.get(desc_col, ""))
            block_cell = normalize_text(row.get(block_col, ""))
            notes = normalize_text(row.get(note_col, ""))
            if not desc or not block_cell:
                continue
            if "no imaging" in desc.lower():
                continue

            blocks = extract_blocks(block_cell)
            if not blocks:
                continue

            meta = parse_jamie_wide_description(desc, notes, mouse)
            for block in blocks:
                out.append(
                    {
                        "mouse": mouse,
                        "phase": current_phase,
                        "date": current_date,
                        "session": current_session,
                        "block": block,
                        "protocol": meta["protocol"],
                        "protocol_raw": meta["protocol_raw"],
                        "fiber_placement": meta["fiber_placement"],
                        "dbs_placement": meta["dbs_placement"],
                        "placement_summary": meta["placement_summary"],
                        "imaging_side": meta["imaging_side"],
                        "frequency_hz": meta["frequency_hz"],
                        "frequency_raw": meta["frequency_raw"],
                        "amplitude_uA": meta["amplitude_uA"],
                        "amplitude_raw": meta["amplitude_raw"],
                        "pulse_width_us": meta["pulse_width_us"],
                        "pulse_width_raw": meta["pulse_width_raw"],
                        "exposure_ms": meta["exposure_ms"],
                        "stimulation_time_s": meta["stimulation_time_s"],
                        "stimulation_time_raw": meta["stimulation_time_raw"],
                        "notes": meta["notes"] if meta["notes"] else "?",
                        "source_sheet": mouse,
                        "source_row": str(row.get(-1, "?")),
                    }
                )

    return out


def fill_from_data(entries: list[dict[str, str]]) -> list[dict[str, str]]:
    for e in entries:
        mouse = e["mouse"]
        date = e["date"]
        block = e["block"]

        e["phase"] = normalize_phase_label(e.get("phase", "?"))

        if not e.get("imaging_side"):
            e["imaging_side"] = infer_imaging_side_from_text(e.get('protocol_raw', ''))
        e["imaging_side"] = IMAGING_SIDE_OVERRIDES.get((mouse, date, block), e["imaging_side"])
        apply_vinnie1_imaging_side(e)

        if is_jamie8_tail_baseline(e):
            e["exposure_ms"] = "2"

        if e["exposure_ms"] == "?":
            e["exposure_ms"] = infer_exposure_ms(mouse, date, block)
        if e["exposure_ms"] == "?":
            e["exposure_ms"] = infer_default_exposure_from_text(e)
        if e["exposure_ms"] == "?":
            e["exposure_ms"] = infer_exposure_ms_from_trace_count(mouse, date, block, e.get("stimulation_time_s", "?"))
        if e["exposure_ms"] == "?" and mouse in {"Vinnie1", "Vinnie2"}:
            e["exposure_ms"] = "1"

        protocol = str(e["protocol"]).lower()
        if (
            e["pulse_width_us"] == "?"
            and "pulsewidth analysis" not in protocol
            and protocol not in {"baseline", "amplitude ramping", "frequency ramping"}
        ):
            try:
                freq = float(e["frequency_hz"])
            except (TypeError, ValueError):
                freq = np.nan
            if np.isfinite(freq) and freq > 0:
                e["pulse_width_us"] = "100"
                if e["pulse_width_raw"] == "?":
                    e["pulse_width_raw"] = "default 100"

        if e["stimulation_time_s"] == "?" and protocol not in {"baseline", "amplitude ramping", "frequency ramping"} and "pulsewidth analysis" not in protocol:
            e["stimulation_time_s"] = infer_stim_duration_s(mouse, date, block)
            if e["stimulation_time_raw"] == "?":
                e["stimulation_time_raw"] = e["stimulation_time_s"]

        if e["mouse"] in ("Jamie8", "Jamie10") and e["amplitude_uA"] == "?" and e["amplitude_raw"] == "TEED":
            try:
                freq = float(e["frequency_hz"])
                pw = float(e["pulse_width_us"])
            except (TypeError, ValueError):
                continue
            e["amplitude_uA"] = infer_teed_amplitude(freq, pw)

        if (
            e["amplitude_uA"] == "?"
            and not is_mixed_cortex_entry(e)
            and str(e["protocol"]).lower() not in {"baseline", "amplitude ramping", "frequency ramping"}
        ):
            try:
                freq = float(e["frequency_hz"])
                pw = float(e["pulse_width_us"])
            except (TypeError, ValueError):
                freq = np.nan
                pw = np.nan
            if np.isfinite(freq) and np.isfinite(pw) and freq > 0 and pw > 0:
                e["amplitude_uA"] = infer_teed_amplitude(freq, pw)
                if e["amplitude_raw"] == "?":
                    e["amplitude_raw"] = "TEED"

        special_amp = infer_special_amplitude_category(e)
        if special_amp is not None:
            e["amplitude_uA"] = special_amp
            e["amplitude_raw"] = special_amp

        for key in ["frequency_hz", "amplitude_uA", "pulse_width_us", "exposure_ms", "stimulation_time_s"]:
            e[key] = normalize_numeric_text(e.get(key, "?"))

    return entries


def build_entries() -> list[dict[str, str]]:
    with zipfile.ZipFile(WORKBOOK_PATH) as zf:
        shared = parse_shared_strings(zf)
        targets = workbook_sheet_targets(zf)

        all_entries: list[dict[str, str]] = []
        for mouse in MICE:
            rows = read_sheet_rows(zf, targets[mouse], shared)
            if mouse == "Jamie6":
                all_entries.extend(parse_jamie6_rows(rows))
            else:
                all_entries.extend(parse_jamie_wide_rows(rows, mouse))

    return fill_from_data(all_entries)


def summary_row(entry: dict[str, str]) -> list[str]:
    return [
        entry.get("mouse", "?"),
        entry.get("phase", "?"),
        entry.get("date", "?"),
        entry.get("block", "?"),
        entry.get("protocol_raw", "?"),
        entry.get("imaging_side", "?"),
        entry.get("frequency_hz", "?"),
        entry.get("amplitude_uA", "?"),
        entry.get("pulse_width_us", "?"),
        entry.get("exposure_ms", "?"),
        entry.get("stimulation_time_s", "?"),
    ]


def exposure_groups(entries: list[dict[str, str]]) -> list[tuple[str, list[dict[str, str]]]]:
    def mouse_rank(mouse: str) -> int:
        try:
            return SUMMARY_MOUSE_ORDER.index(mouse)
        except ValueError:
            return len(SUMMARY_MOUSE_ORDER)

    def sort_key(e: dict[str, str]):
        special_tail = 1 if is_jamie8_tail_baseline(e) else 0
        return (
            mouse_rank(e.get("mouse", "")),
            special_tail,
            parse_date_for_sort(e.get("date", "?")),
            block_sort_value(e.get("block", "?")),
        )

    mixed_special = [e for e in entries if is_mixed_cortex_entry(e)]
    regular = [e for e in entries if not is_mixed_cortex_entry(e)]
    filtered = sorted(regular, key=sort_key)
    groups: list[tuple[str, list[dict[str, str]]]] = []
    for value, label in [("1", "1 ms"), ("2", "2 ms")]:
        group_entries = [e for e in filtered if str(e.get("exposure_ms", "?")) == value]
        if group_entries:
            groups.append((label, group_entries))

    other_values = []
    for e in filtered:
        exp = str(e.get("exposure_ms", "?"))
        if exp not in {"1", "2"} and exp not in other_values:
            other_values.append(exp)

    if "?" in other_values:
        other_values = ["?"] + [v for v in other_values if v != "?"]

    for exp in other_values:
        group_entries = [e for e in filtered if str(e.get("exposure_ms", "?")) == exp]
        if not group_entries:
            continue
        label = f"Other ({exp} ms)" if exp != "?" else "Other (?)"
        groups.append((label, group_entries))

    if mixed_special:
        groups.append(("Mixed Cortex / Special", sorted(mixed_special, key=sort_key)))

    return groups


def save_summary_xlsx(entries: list[dict[str, str]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = ["Mouse", "Phase", "Date", "Block", "Protocol", "ImagingSide", "Freq_Hz", "Amp_uA", "PW_us", "Exposure_ms", "StimTime_s"]
    section_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    mouse_fills = {
        "Jamie6": PatternFill("solid", fgColor="FFF2CC"),
        "Jamie8": PatternFill("solid", fgColor="E2F0D9"),
        "Jamie10": PatternFill("solid", fgColor="DDEBF7"),
        "Jamie11": PatternFill("solid", fgColor="FCE4D6"),
        "Jamie12": PatternFill("solid", fgColor="E4DFEC"),
        "Vinnie1": PatternFill("solid", fgColor="DAEEF3"),
        "Vinnie2": PatternFill("solid", fgColor="EBF1DE"),
    }
    thin = Side(style="thin", color="D0D0D0")
    medium = Side(style="medium", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_idx = 1
    for section_label, section_entries in exposure_groups(entries):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        cell = ws.cell(row_idx, 1, section_label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        for c, header in enumerate(headers, start=1):
            cell = ws.cell(row_idx, c, header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        last_mouse = None
        for entry in section_entries:
            values = summary_row(entry)
            row_mouse = entry.get("mouse", "")
            row_border = border
            if last_mouse is not None and row_mouse != last_mouse:
                row_border = Border(left=thin, right=thin, top=medium, bottom=thin)
            for c, value in enumerate(values, start=1):
                cell = ws.cell(row_idx, c, value)
                cell.border = row_border
                cell.fill = mouse_fills.get(entry.get("mouse", ""), PatternFill(fill_type=None))
                if c == 4:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                elif c in {6, 7, 8, 9, 10, 11}:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                if c in {1, 2, 3}:
                    cell.font = Font(bold=True)
            ws.row_dimensions[row_idx].height = 30
            last_mouse = row_mouse
            row_idx += 1
        row_idx += 1

    widths = {
        "A": 12,
        "B": 10,
        "C": 11,
        "D": 8,
        "E": 82,
        "F": 12,
        "G": 12,
        "H": 10,
        "I": 12,
        "J": 12,
        "K": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "D3"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    wb.save(path)


def main() -> None:
    entries = build_entries()
    fieldnames = [
        "mouse",
        "phase",
        "date",
        "session",
        "block",
        "protocol",
        "protocol_raw",
        "imaging_side",
        "fiber_placement",
        "dbs_placement",
        "placement_summary",
        "frequency_hz",
        "frequency_raw",
        "amplitude_uA",
        "amplitude_raw",
        "pulse_width_us",
        "pulse_width_raw",
        "exposure_ms",
        "stimulation_time_s",
        "stimulation_time_raw",
        "notes",
        "source_sheet",
        "source_row",
    ]
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(entries)

    saved_summary_path = SUMMARY_XLSX
    try:
        save_summary_xlsx(entries, SUMMARY_XLSX)
    except PermissionError:
        saved_summary_path = SUMMARY_XLSX.with_name(f"{SUMMARY_XLSX.stem}_updated{SUMMARY_XLSX.suffix}")
        save_summary_xlsx(entries, saved_summary_path)

    print(f"[SAVED] {OUTPUT_CSV}")
    print(f"[SAVED] {saved_summary_path}")
    print(f"[ROWS] {len(entries)}")


if __name__ == "__main__":
    main()
